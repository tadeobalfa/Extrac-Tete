# process_bbva_v3.py — Parser BBVA (AR) con pdfplumber
# Ajustes:
# 1) Para renglones tipo "IMP.LEY ...." sin importes explícitos, si hay saldo nuevo, usar
#    la diferencia de saldo respecto del saldo anterior como importe (crédito/débito).
# 2) Corte de cuenta robusto: frena al detectar frases que indican otras secciones
#    ("ENVIADAS ACEPTADAS", "LE INFORMAMOS QUE", "INVERSIONES EN BONOS/FONDOS/ACCIONES", "LEGALES Y AVISOS", etc.).

from __future__ import annotations
import io, math, re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd
import pdfplumber

# ---------------- Regex base ----------------
DATE_RE = re.compile(r"\b([0-3]?\d)[/-]([01]?\d)(?:[/-](\d{2,4}))?\b")
AMT_IN_TEXT_RE = re.compile(r"[-+]?\s*\$?\s*(?:\d{1,3}(?:\.\d{3})+|\d+)(?:,\d{2})(?:-)?")
AMOUNT_RE_STRICT = re.compile(r"^\s*\$?\s*[-+]?(?:\d{1,3}(?:\.\d{3})+|\d+)(?:,\d{2})?\s*-?\s*$")
CURRENCY_STRIP = re.compile(r"[\s\$]")
HEADER_YEAR_RE = re.compile(r"\b(20\d{2}|19\d{2})\b", re.IGNORECASE)

# Encabezados de CUENTA
ACC_HEADER_RE = re.compile(r"\bCC\s+(U\$S|\$)\s+([0-9][0-9\.\-\s]*)[/-]\s*([0-9]+)\b", re.IGNORECASE)

MOV_HEADER_RE = re.compile(r"\bMOVIMIENTOS\s+EN\s+CUENTAS\b", re.IGNORECASE)
END_ACC_RE = re.compile(r"\b(SALDO\s+AL|TOTAL\s+MOVIMIENTOS)\b", re.IGNORECASE)
SALDO_ANT_RE = re.compile(r"\bSALDO\s+ANTERIOR\b", re.IGNORECASE)

# Marcas que INDICAN fin real de la grilla de movimientos aunque la página continúe
HARD_END_MARKERS = tuple(s.upper() for s in [
    "ENVIADAS ACEPTADAS",
    "LE INFORMAMOS QUE",
    "INVERSIONES EN BONOS",
    "INVERSIONES EN FONDOS",
    "INVERSIONES EN ACCIONES",
    "LEGALES Y AVISOS",
    "DETALLE DE IMPUESTO",
    "TOTAL SALDOS DISPONIBLES",
])

# Limpieza de descripción
NOISE_RE = re.compile(r"\(cid:\d+\)|cid:\d+", re.IGNORECASE)
LEADING_DUP_DATE_RE = re.compile(r"^\s*\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?\s+")
BLOCKLIST_DESC = tuple(s.upper() for s in [
    "WWW.BBVA","LOS DEPÓSITOS EN PESOS","PÁGINA","CUENTAS Y PAQUETES",
    "CUENTAS Y PAQUETE","OCASA","R.N.P.S.P."
])

IGNORE_LINE_MARKERS = (
    "SIN MOVIMIENTOS",
    "RECIBIDAS (INFORMACIÓN AL", "RECIBIDAS (INFORMACION AL",
    "CONSULTAS Y RECLAMOS",
)

try:
    from sklearn.cluster import KMeans
    HAS_SKLEARN = True
except Exception:
    HAS_SKLEARN = False


@dataclass
class PageColumns:
    cut1: float
    cut2: float
    left_border: float


# ---------------- Utilidades ----------------
def _clean_amount(txt: str) -> Optional[float]:
    if not txt: return None
    s = CURRENCY_STRIP.sub("", txt).replace(".", "").replace(",", ".").strip()
    neg = False
    if s.endswith("-"): neg = True; s = s[:-1].strip()
    if s.startswith("-"): neg = True; s = s[1:].strip()
    if s.startswith("+"): s = s[1:].strip()
    try:
        val = float(s)
    except ValueError:
        return None
    return -val if neg else val

def _is_amount_strict(t: str) -> bool:
    return bool(AMOUNT_RE_STRICT.match(t or ""))

def _is_dateish(t: str) -> bool:
    return bool(DATE_RE.search(t or ""))

def _parse_date(txt: str, default_year: int, last_date: Optional[pd.Timestamp]) -> Optional[pd.Timestamp]:
    m = DATE_RE.search(txt or "")
    if not m: return None
    dd, mm, yy = m.group(1), m.group(2), m.group(3)
    day = int(dd); month = int(mm)
    if yy:
        y = int(yy)
        if y < 100: y = y + (2000 if y <= 69 else 1900)
    else:
        if last_date is None: y = default_year
        else:
            y = last_date.year + 1 if (month < last_date.month and (last_date.month - month) >= 6) else last_date.year
    try:
        return pd.Timestamp(year=y, month=month, day=day)
    except Exception:
        return None

def _clean_description(desc: str) -> str:
    if not desc: return ""
    d = NOISE_RE.sub("", desc)
    d = LEADING_DUP_DATE_RE.sub("", d)
    d = re.sub(r"\s{2,}", " ", d).strip(" -—•:")
    if any(bad in d.upper() for bad in BLOCKLIST_DESC): return ""
    return d

def _normalize_dc(debito: float, credito: float) -> Tuple[float, float]:
    debito = float(debito or 0.0); credito = float(credito or 0.0)
    if debito < 0 and credito == 0: credito = abs(debito); debito = 0.0
    elif credito < 0 and debito == 0: debito = abs(credito); credito = 0.0
    if abs(debito) > 0 and abs(credito) > 0:
        if math.isclose(abs(debito), abs(credito), abs_tol=0.005): credito = abs(credito); debito = 0.0
        else:
            if abs(debito) > abs(credito): credito = 0.0
            else: debito = 0.0
    debito  = abs(debito)  if abs(debito)  >= 0.004 else 0.0
    credito = abs(credito) if abs(credito) >= 0.004 else 0.0
    return debito, credito

def _group_words_by_line(words: List[dict], y_tol: float = 2.0) -> List[List[dict]]:
    if not words: return []
    ws = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines: List[List[dict]] = []; cur: List[dict] = []; last = None
    for w in ws:
        if last is None or abs(w["top"] - last) > y_tol:
            if cur: lines.append(sorted(cur, key=lambda ww: ww["x0"]))
            cur = [w]; last = w["top"]
        else:
            cur.append(w); last = (last + w["top"]) / 2.0
    if cur: lines.append(sorted(cur, key=lambda ww: ww["x0"]))
    return lines

def _discover_columns(words: List[dict], page_w: float) -> Optional[PageColumns]:
    """
    Detecta columnas Débito / Crédito / Saldo.
    Primero intenta por encabezados textuales.
    Si no puede, cae a clustering de importes.
    """

    # -------- Intento 1: usar encabezados --------
    header_deb = None
    header_cred = None
    header_saldo = None

    for w in words:
        txt = (w.get("text", "") or "").strip().upper()
        xc = (w["x0"] + w["x1"]) / 2.0

        if txt in {"DEBITO", "DÉBITO"}:
            header_deb = xc
        elif txt in {"CREDITO", "CRÉDITO"}:
            header_cred = xc
        elif txt == "SALDO":
            header_saldo = xc

    if header_deb and header_cred and header_saldo:
        return PageColumns(
            cut1=(header_deb + header_cred) / 2.0,
            cut2=(header_cred + header_saldo) / 2.0,
            left_border=max(0, header_deb - 120),
        )

    # -------- Intento 2: clustering por importes --------
    xs = []
    for w in words:
        t = w.get("text", "")
        if not _is_amount_strict(t):
            continue
        xc = (w["x0"] + w["x1"]) / 2.0

        # antes estaba en 0.45 y podía dejar afuera Débito
        if xc >= page_w * 0.30:
            xs.append(xc)

    if len(xs) < 3:
        return None

    centers = None
    try:
        if HAS_SKLEARN:
            X = np.array(xs).reshape(-1, 1)
            km = KMeans(
                n_clusters=3,
                n_init="auto" if hasattr(KMeans, "n_init") else 10,
                random_state=0
            )
            km.fit(X)
            centers = sorted([c[0] for c in km.cluster_centers_])
    except Exception:
        centers = None

    if centers is None:
        xs2 = sorted(xs)
        groups = [[xs2[0]]]
        for x in xs2[1:]:
            if abs(x - groups[-1][-1]) <= 12:
                groups[-1].append(x)
            else:
                groups.append([x])

        groups = sorted(groups, key=lambda g: (len(g), np.median(g)), reverse=True)[:3]
        centers = sorted([float(np.median(g)) for g in groups])

    if len(centers) != 3:
        return None

    debit_c, credit_c, saldo_c = centers

    return PageColumns(
        cut1=(debit_c + credit_c) / 2.0,
        cut2=(credit_c + saldo_c) / 2.0,
        left_border=max(0, debit_c - 120),
    )

def _take_last_amount_from_tokens(tokens: List[str]) -> float:
    last: Optional[float] = None
    for tok in tokens:
        for m in AMT_IN_TEXT_RE.findall(tok):
            val = _clean_amount(m)
            if val is not None: last = val
    return float(last) if last is not None else 0.0

def _fmt_sheet_name(currency: str, left: str, right: str) -> str:
    left = left.replace(".", "-").replace(" ", "")
    left = re.sub(r"[^\d\-]", "", left)
    return f"CC {currency} {left}-{right}".replace("/", "-")


# --------------- Parser principal ----------------
def parse_bbva_pdf(source: Union[str, bytes, io.BytesIO]) -> Dict[str, pd.DataFrame]:
    pdf = pdfplumber.open(io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source)

    sheets: Dict[str, pd.DataFrame] = {}
    current_account: Optional[str] = None
    movement_section = False
    skip_until_header = True  # hasta ver el primer "CC ..."
    prev_saldo_map: Dict[str, Optional[float]] = {}  # saldo previo por cuenta

    try:
        last_date: Optional[pd.Timestamp] = None
        default_year: Optional[int] = None

        for page in pdf.pages:
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False, x_tolerance=1, y_tolerance=1) or []
            if not words:
                raw = page.extract_text() or ""
                words = [{"text": t, "x0": 0, "x1": 0, "top": i*10, "bottom": i*10+9} for i, t in enumerate(raw.splitlines())]

            page_w = page.width
            lines = _group_words_by_line(words, y_tol=2.0)
            lines_text = [" ".join([w["text"] for w in ln]) for ln in lines]

            # Año
            if default_year is None:
                years = [int(y) for t in lines_text[:25] for y in HEADER_YEAR_RE.findall(t)]
                default_year = max(set(years), key=years.count) if years else (last_date.year if last_date is not None else pd.Timestamp.today().year)

            # ¿aparece la sección?
            if any(MOV_HEADER_RE.search(t) for t in lines_text):
                movement_section = True
                current_account = None
                skip_until_header = True

            if not movement_section:
                continue

            # ¿encabezado de cuenta?
            if skip_until_header:
                found = False
                for t in lines_text:
                    m = ACC_HEADER_RE.search(t)
                    if m:
                        currency, left, right = m.group(1), m.group(2), m.group(3)
                        current_account = _fmt_sheet_name(currency, left, right)
                        sheets.setdefault(current_account, pd.DataFrame(columns=["Fecha","Descripción","Débito","Crédito","Saldo"]))
                        prev_saldo_map.setdefault(current_account, None)
                        skip_until_header = False
                        found = True
                        break
                if not found:
                    continue

            cols = _discover_columns(words, page_w)
            print("DEBUG BBVA COLS:", current_account, "page_w=", page_w,
      		  "left_border=", None if cols is None else round(cols.left_border, 2),
      		  "cut1=", None if cols is None else round(cols.cut1, 2),
      		  "cut2=", None if cols is None else round(cols.cut2, 2))
            if not cols:
                continue

            i = 0
            while i < len(lines):
                ln = lines[i]
                text_line = " ".join([w["text"] for w in ln]).strip()
                up = text_line.upper()

                # Fin de cuenta (marcadores "oficiales")
                if END_ACC_RE.search(up):
                    current_account = None
                    skip_until_header = True
                    i += 1
                    continue

                # Cortes "duros" si empiezan bloques no bancarios
                if any(mark in up for mark in HARD_END_MARKERS):
                    current_account = None
                    skip_until_header = True
                    i += 1
                    continue

                # Nuevo encabezado de cuenta
                mhead = ACC_HEADER_RE.search(text_line)
                if mhead:
                    currency, left, right = mhead.group(1), mhead.group(2), mhead.group(3)
                    current_account = _fmt_sheet_name(currency, left, right)
                    sheets.setdefault(current_account, pd.DataFrame(columns=["Fecha","Descripción","Débito","Crédito","Saldo"]))
                    prev_saldo_map.setdefault(current_account, None)
                    skip_until_header = False
                    i += 1
                    continue

                if current_account is None:
                    i += 1
                    continue

                if SALDO_ANT_RE.search(up):
                    # no registrar el saldo anterior como movimiento
                    i += 1
                    continue

                if not _is_dateish(text_line):
                    i += 1
                    continue

                # Fecha del movimiento
                date_token_text = None
                for w in ln[:5]:
                    if _is_dateish(w["text"]):
                        date_token_text = w["text"]; break
                date_obj = _parse_date(date_token_text or text_line, default_year, last_date)

                # Descripción e importes (puede ocupar varias líneas)
                desc_parts: List[str] = []
                deb_tokens: List[str] = []
                cred_tokens: List[str] = []
                saldo_tokens: List[str] = []
                end_reached = False

                j = i
                while j < len(lines):
                    ln2 = lines[j]
                    txt2 = " ".join([w["text"] for w in ln2]).strip()
                    up2 = txt2.upper()

                    if j != i and _is_dateish(txt2): break
                    if END_ACC_RE.search(up2): end_reached = True; break
                    if any(mark in up2 for mark in HARD_END_MARKERS): end_reached = True; break
                    if ACC_HEADER_RE.search(txt2): break

                    for w in ln2:
                        t = w["text"]; xc = (w["x0"] + w["x1"]) / 2.0

                        if xc < cols.left_border:
                            if not _is_dateish(t):
                                desc_parts.append(t)
                        elif cols.left_border <= xc < cols.cut1:
                            deb_tokens.append(t)
                        elif cols.cut1 <= xc < cols.cut2:
                            cred_tokens.append(t)
                        else:
                            saldo_tokens.append(t)
                    j += 1

                deb = _take_last_amount_from_tokens(deb_tokens)
                cred = _take_last_amount_from_tokens(cred_tokens)
                saldo_text = _take_last_amount_from_tokens(saldo_tokens)
                saldo_val = None if saldo_text == 0.0 else float(saldo_text)

                deb, cred = _normalize_dc(deb, cred)
                descripcion_raw = " ".join(desc_parts).strip() or DATE_RE.sub("", text_line, count=1).strip()
                descripcion = _clean_description(descripcion_raw)

                # Si hay importe al final de la descripción, úsalo solo si deb/cred == 0
                if descripcion and abs(deb) < 0.004 and abs(cred) < 0.004:
                    mtrail = re.search(r"(?:\s|^)([-+]?\s*\$?\s*(?:\d{1,3}(?:\.\d{3})+|\d+)(?:,\d{2})(?:-)?)[\s]*$", descripcion)
                    if mtrail:
                        val = _clean_amount(mtrail.group(1))
                        if val is not None:
                            if val < 0: deb = abs(val); cred = 0.0
                            else: cred = val; deb = 0.0
                            descripcion = descripcion[:mtrail.start()].rstrip(" -—•:")

                # ------ FIX SUAVE: delta de saldo para "IMP.LEY ..." si no hay importes explícitos ------
                if (abs(deb) < 0.004 and abs(cred) < 0.004) and saldo_val is not None and descripcion:
                    if "IMP.LEY" in descripcion.upper():
                        prev = prev_saldo_map.get(current_account)
                        if prev is not None:
                            diff = saldo_val - prev
                            if abs(diff) >= 0.004:
                                if diff > 0: cred = diff
                                else: deb = abs(diff)

                if not descripcion:
                    i = j
                    if end_reached:
                        current_account = None
                        skip_until_header = True
                    continue
                if any(marker in descripcion.upper() for marker in IGNORE_LINE_MARKERS):
                    i = j
                    if end_reached:
                        current_account = None
                        skip_until_header = True
                    continue

                df = sheets[current_account]
                df.loc[len(df)] = [date_obj, descripcion, deb, cred, saldo_val]

                # actualizar saldo previo de la cuenta
                if saldo_val is not None:
                    prev_saldo_map[current_account] = float(saldo_val)

                last_date = date_obj
                i = j

                if end_reached:
                    current_account = None
                    skip_until_header = True

    finally:
        pdf.close()

    # Normalización final
    for acc, df in list(sheets.items()):
        if df.empty: continue
        df["Fecha"]   = pd.to_datetime(df["Fecha"], errors="coerce")
        df["Débito"]  = pd.to_numeric(df["Débito"],  errors="coerce").fillna(0.0)
        df["Crédito"] = pd.to_numeric(df["Crédito"], errors="coerce").fillna(0.0)
        df["Saldo"]   = pd.to_numeric(df["Saldo"],   errors="coerce")
        if df["Saldo"].isna().any():
            fk = df["Saldo"].first_valid_index()
            if fk is not None:
                running = float(df.loc[fk, "Saldo"])
                for k in range(fk + 1, len(df)):
                    if pd.isna(df.at[k, "Saldo"]):
                        running = running - float(df.at[k, "Débito"]) + float(df.at[k, "Crédito"])
                        df.at[k, "Saldo"] = running
        sheets[acc] = df

    return sheets


# ---------------- Excel helpers / API ----------------
_SHEET_BAD_CHARS_RE = re.compile(r'[:\\/?*\[\]]')

def _unique_safe_name(name: str, used: set[str]) -> str:
    base = _SHEET_BAD_CHARS_RE.sub("-", (name or "Hoja").strip()).replace("/", "-")
    base = base[:31] if len(base) > 31 else base
    if not base: base = "Hoja"
    cand = base; n = 2
    while cand in used:
        suf = f" ({n})"
        cand = base[:31-len(suf)] + suf if len(base) + len(suf) > 31 else base + suf
        n += 1
    used.add(cand); return cand

def to_excel(sheets: Dict[str, pd.DataFrame], output_path: str, order: Optional[List[str]] = None) -> str:
    from openpyxl.utils import get_column_letter
    used: set[str] = set()
    items = list(sheets.items()) if not order else [(k, sheets[k]) for k in order if k in sheets]
    with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="DD/MM/YYYY", date_format="DD/MM/YYYY") as w:
        for name, df in items:
            if df is None or df.empty: continue
            safe = _unique_safe_name(name, used)
            df2 = df.copy()
            df2["Fecha"] = pd.to_datetime(df2["Fecha"], errors="coerce")
            for c in ("Débito","Crédito","Saldo"):
                df2[c] = pd.to_numeric(df2[c], errors="coerce")
            df2.to_excel(w, index=False, sheet_name=safe)
            ws = w.sheets[safe]
            # negrita encabezados
            for cell in ws[1]:
                try:
                    cell.font = cell.font.copy(bold=True)
                except Exception:
                    pass
            headers = [c.value for c in ws[1]]
            cmap = {h:i+1 for i,h in enumerate(headers)}
            if "Fecha" in cmap:
                c = cmap["Fecha"]
                for r in range(2, ws.max_row+1): ws.cell(row=r, column=c).number_format = "DD/MM/YYYY"
            for h in ("Débito","Crédito","Saldo"):
                if h in cmap:
                    c = cmap[h]
                    for r in range(2, ws.max_row+1): ws.cell(row=r, column=c).number_format = "#,##0.00"
            widths = {"Fecha":12,"Descripción":70,"Débito":15,"Crédito":15,"Saldo":15}
            for h,wth in widths.items():
                if h in cmap:
                    ws.column_dimensions[get_column_letter(cmap[h])].width = wth
    return output_path

def process_bbva(input_sources: List[Union[str, bytes, io.BytesIO]], output_excel_path: Optional[str] = None) -> Dict[str, pd.DataFrame]:
    all_sheets: Dict[str, pd.DataFrame] = {}
    order: List[str] = []
    for src in input_sources:
        sheets = parse_bbva_pdf(src)
        for k, df in sheets.items():
            if k not in all_sheets:
                all_sheets[k] = df; order.append(k)
            else:
                all_sheets[k] = pd.concat([all_sheets[k], df], ignore_index=True)
    if output_excel_path:
        to_excel(all_sheets, output_excel_path, order=order)
    return all_sheets

def _cli():
    import argparse
    p = argparse.ArgumentParser(description="Parser de extractos BBVA (AR) → Excel (V3 ajustes suaves)")
    p.add_argument("pdfs", nargs="+", help="Ruta(s) BBVA.pdf")
    p.add_argument("-o","--output", default="BBVA.xlsx")
    a = p.parse_args()
    sheets = process_bbva(a.pdfs, output_excel_path=a.output)
    print("OK:", a.output, "| Hojas:", ", ".join(sheets))

if __name__ == "__main__":
    _cli()
